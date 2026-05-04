#!/usr/bin/env python3
"""Build the DCFS ILC pilot baseline survey questions as a single-sheet Excel file.

Reads the question content baked into this script (sourced from
guide/appendix-d-templates.md Templates 1, 2, 3 + DCFS adaptations from
pilot-baseline-survey-questions-v20260423.md).

Writes: pilot-baseline-survey-questions-v20260427.xlsx
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

OUT = os.path.join(os.path.dirname(__file__), 'pilot-baseline-survey-questions-v20260427.xlsx')

LIKERT_5 = '1-5 scale (1=Strongly disagree, 2=Disagree, 3=Neutral, 4=Agree, 5=Strongly agree)'
LIKERT_5_SKILL = '1-5 scale (1=No experience, 2=Aware but not practiced, 3=Can do with guidance, 4=Comfortable independently, 5=Could teach others)'

# (#, section, qid, question, type, scale_options, dcfs_note)
ROWS = [
    # ---- Template 1: AI Readiness Assessment ----
    # Section A
    (1, 'Template 1 / Section A: Current AI Experience', 'A1',
     'Have you used any AI-assisted development or productivity tools professionally?',
     'Single-select',
     'Yes regularly (daily/weekly) | Yes occasionally (few times/month) | Yes briefly or trial | No but used personally | No, have not used',
     ''),
    (2, 'Template 1 / Section A: Current AI Experience', 'A1a',
     'In your current ILC pilot work (last 30 days), how often have you used AI tools (Copilot, Rovo, M365 Copilot, D365 Copilot)?',
     'Single-select',
     'Daily | Several times a week | A few times a month | Tried once or twice | Not yet — no access | Not yet — access available, haven\'t started',
     'DCFS adaptation: pins prior-usage baseline to current ILC work specifically — critical for interpreting post-training delta'),
    (3, 'Template 1 / Section A: Current AI Experience', 'A2',
     'Which AI tools have you used? (Select all that apply)',
     'Multi-select',
     'GitHub Copilot | Atlassian Rovo / Atlassian Intelligence | Microsoft 365 Copilot | Dynamics 365 Copilot | ChatGPT / GPT-4 | Google Gemini | Claude (Anthropic) | Other code completion tools | Other AI productivity tools (specify) | None',
     'DCFS adaptation: added Microsoft 365 Copilot and Dynamics 365 Copilot to options'),
    (3, 'Template 1 / Section A: Current AI Experience', 'A3',
     'In which SDLC processes have you applied AI tools? (Select all that apply)',
     'Multi-select',
     'Code writing/generation | User stories / requirements | Test case creation | Code review / PR analysis | Documentation | Bug triage / defect analysis | Sprint planning / estimation | None',
     ''),
    (4, 'Template 1 / Section A: Current AI Experience', 'A4',
     'How would you describe the outcome of your AI tool usage?',
     'Single-select',
     'Significantly improved productivity | Somewhat improved productivity | Neutral | Made things slower or more complicated | Not applicable',
     ''),
    (5, 'Template 1 / Section A: Current AI Experience', 'A5',
     'Have you received formal AI tools training?',
     'Single-select',
     'Yes, comprehensive (multi-day / certification) | Yes, introductory (workshop / webinar) | Self-taught only | No training',
     ''),
    # Section B - Skills
    (6, 'Template 1 / Section B: Skills Self-Assessment', 'B1',
     'Writing effective prompts for AI tools', 'Likert 1-5', LIKERT_5_SKILL, ''),
    (7, 'Template 1 / Section B: Skills Self-Assessment', 'B2',
     'Reviewing AI-generated code for correctness/security', 'Likert 1-5', LIKERT_5_SKILL, ''),
    (8, 'Template 1 / Section B: Skills Self-Assessment', 'B3',
     'Using AI for user story writing / requirements', 'Likert 1-5', LIKERT_5_SKILL, ''),
    (9, 'Template 1 / Section B: Skills Self-Assessment', 'B4',
     'Using AI for test case generation', 'Likert 1-5', LIKERT_5_SKILL, ''),
    (10, 'Template 1 / Section B: Skills Self-Assessment', 'B5',
     'Integrating AI output into existing SDLC workflows', 'Likert 1-5', LIKERT_5_SKILL, ''),
    (11, 'Template 1 / Section B: Skills Self-Assessment', 'B6',
     'Identifying wrong/incomplete AI suggestions', 'Likert 1-5', LIKERT_5_SKILL, ''),
    # Section C - Process
    (12, 'Template 1 / Section C: Process Readiness', 'C1',
     'Our team has well-defined, documented SDLC processes', 'Likert 1-5', LIKERT_5, ''),
    (13, 'Template 1 / Section C: Process Readiness', 'C2',
     'Our team regularly reviews and improves workflows', 'Likert 1-5', LIKERT_5, ''),
    (14, 'Template 1 / Section C: Process Readiness', 'C3',
     'Our team has clear code review standards', 'Likert 1-5', LIKERT_5, ''),
    (15, 'Template 1 / Section C: Process Readiness', 'C4',
     'Our team can adopt new tools without disrupting work', 'Likert 1-5', LIKERT_5, ''),
    (16, 'Template 1 / Section C: Process Readiness', 'C5',
     'Our Scrum Master supports experimentation', 'Likert 1-5', LIKERT_5, ''),
    # Section D - Attitudes
    (17, 'Template 1 / Section D: Attitudes & Concerns', 'D1',
     'AI tools can meaningfully improve my work quality', 'Likert 1-5', LIKERT_5, ''),
    (18, 'Template 1 / Section D: Attitudes & Concerns', 'D2',
     'AI tools will augment my role, not replace it', 'Likert 1-5', LIKERT_5, ''),
    (19, 'Template 1 / Section D: Attitudes & Concerns', 'D3',
     'I am willing to invest time learning AI tools', 'Likert 1-5', LIKERT_5, ''),
    (20, 'Template 1 / Section D: Attitudes & Concerns', 'D4',
     'Structured AI adoption is better than ad-hoc', 'Likert 1-5', LIKERT_5, ''),
    (21, 'Template 1 / Section D: Attitudes & Concerns', 'D5',
     'What is your biggest concern about AI tool adoption? (Including concerns about DCFS data boundaries -- CANTS, CCWIS, PII)',
     'Open text', 'Free response',
     'DCFS adaptation: added prompt about DCFS data boundaries'),
    (22, 'Template 1 / Section D: Attitudes & Concerns', 'D6',
     'What SDLC process could AI most improve for your team, and why?',
     'Open text', 'Free response', ''),
    # Section E - Infrastructure (team-level, completed by Scrum Master / Lead)
    (23, 'Template 1 / Section E: Infrastructure (team-level, completed by Scrum Master/Lead)', 'E1',
     'GitHub Copilot access',
     'Single-select',
     'All members | Some | Requested | No | Unsure',
     ''),
    (24, 'Template 1 / Section E: Infrastructure (team-level, completed by Scrum Master/Lead)', 'E2',
     'Atlassian Rovo access',
     'Single-select',
     'Enabled | Partial | Available | No | Unsure',
     ''),
    (25, 'Template 1 / Section E: Infrastructure (team-level, completed by Scrum Master/Lead)', 'E2a',
     'Microsoft 365 Copilot access',
     'Single-select',
     'All members | Some | Requested | No | Unsure',
     'DCFS adaptation: added M365 Copilot access row'),
    (26, 'Template 1 / Section E: Infrastructure (team-level, completed by Scrum Master/Lead)', 'E2b',
     'Dynamics 365 Copilot access',
     'Single-select',
     'All members | Some | Requested | No | Unsure',
     'DCFS adaptation: added D365 Copilot access row'),
    (27, 'Template 1 / Section E: Infrastructure (team-level, completed by Scrum Master/Lead)', 'E3',
     'Constraints (Select all that apply)',
     'Multi-select',
     'Network/firewall | Data policy | IDE/toolchain | Procurement | No known constraints | Unsure',
     ''),
    # ---- Template 2: SPACE Survey ----
    (28, 'Template 2: SPACE Survey (S - Satisfaction)', 'S1',
     'I am satisfied with my ability to do my job effectively using the tools and processes available to me.',
     'Likert 1-5', LIKERT_5, ''),
    (29, 'Template 2: SPACE Survey (P - Performance)', 'S2',
     'I am confident that the work I produce meets the quality standards expected by my team and stakeholders.',
     'Likert 1-5', LIKERT_5, ''),
    (30, 'Template 2: SPACE Survey (A - Activity)', 'S3',
     'I feel that I am able to complete a meaningful amount of work during a typical sprint.',
     'Likert 1-5', LIKERT_5, ''),
    (31, 'Template 2: SPACE Survey (C - Communication & Collaboration)', 'S4',
     'My team communicates effectively, and I have the information I need to do my work without unnecessary delays or blockers.',
     'Likert 1-5', LIKERT_5, ''),
    (32, 'Template 2: SPACE Survey (E - Efficiency & Flow)', 'S5',
     'I am able to focus on my work without frequent context-switching, unnecessary meetings, or process friction slowing me down.',
     'Likert 1-5', LIKERT_5, ''),
    # ---- Template 3: DevEx Survey ----
    (33, 'Template 3: DevEx Survey (Flow State)', 'X1',
     'I regularly experience periods of deep, uninterrupted focus during my work where I can make meaningful progress on complex tasks.',
     'Likert 1-5', LIKERT_5, ''),
    (34, 'Template 3: DevEx Survey (Feedback Loops)', 'X2',
     'When I complete work (code, stories, test cases), I receive timely feedback that helps me improve and move forward without long waiting periods.',
     'Likert 1-5', LIKERT_5, ''),
    (35, 'Template 3: DevEx Survey (Cognitive Load)', 'X3',
     'The tools, processes, and systems I use are intuitive enough that I can focus on the problem I am solving rather than fighting the tools or processes themselves.',
     'Likert 1-5', LIKERT_5, ''),
    # ---- Optional: DCFS Compliance Awareness ----
    (36, 'Optional Block: DCFS Compliance Awareness (reviewer-confirm to include)', 'CA1',
     'I am confident I know what data I can and cannot enter into an AI prompt on this program.',
     'Likert 1-5', LIKERT_5,
     'DCFS-specific addition. Reviewers: include or exclude.'),
    (37, 'Optional Block: DCFS Compliance Awareness (reviewer-confirm to include)', 'CA2',
     'I know what to do if I accidentally enter PII or case data into an AI prompt.',
     'Likert 1-5', LIKERT_5,
     'DCFS-specific addition. Reviewers: include or exclude.'),
    (38, 'Optional Block: DCFS Compliance Awareness (reviewer-confirm to include)', 'CA3',
     'I understand the DoIT AI Policy requirements that apply to my role.',
     'Likert 1-5', LIKERT_5,
     'DCFS-specific addition. Reviewers: include or exclude.'),
]

wb = Workbook()
ws = wb.active
ws.title = 'Pilot Baseline Survey'

# ---- Header / metadata block ----
ws['A1'] = 'DCFS ILC AI Rollout — Pilot Baseline Survey Questions'
ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
ws['A1'].fill = PatternFill('solid', fgColor='0F3460')
ws.merge_cells('A1:G1')
ws['A1'].alignment = Alignment(horizontal='left', vertical='center')

ws['A2'] = 'Version: v0.1 — 2026-04-27 — Initial draft for review'
ws['A2'].font = Font(italic=True, size=10)
ws.merge_cells('A2:G2')

ws['A3'] = ('Sources: AI Readiness Assessment (Template 1), SPACE Survey (Template 2, Forsgren et al. 2021), '
            'DevEx Survey (Template 3). Total questions: 36 core + 3 optional compliance. '
            'Estimated completion time: 10–15 minutes per respondent.')
ws['A3'].font = Font(size=10)
ws['A3'].alignment = Alignment(wrap_text=True, vertical='top')
ws.merge_cells('A3:G3')
ws.row_dimensions[3].height = 36

# Spacer row
ws.row_dimensions[4].height = 6

# ---- Column headers ----
HEADER_ROW = 5
headers = ['#', 'Section', 'Question ID', 'Question', 'Type', 'Scale / Options', 'DCFS Note / Reviewer Comments']
for col_idx, h in enumerate(headers, start=1):
    cell = ws.cell(row=HEADER_ROW, column=col_idx, value=h)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor='0F3460')
    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

# ---- Data rows ----
border = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0'),
)

for row_idx, row_data in enumerate(ROWS, start=HEADER_ROW + 1):
    seq_num = row_idx - HEADER_ROW  # auto-renumber regardless of hardcoded # in tuple
    for col_idx, value in enumerate(row_data, start=1):
        if col_idx == 1:
            value = seq_num
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        cell.border = border
        # Optional compliance block — light yellow shading to distinguish
        if isinstance(row_data[1], str) and row_data[1].startswith('Optional Block'):
            cell.fill = PatternFill('solid', fgColor='FFF8E1')

# ---- Column widths ----
widths = {1: 5, 2: 32, 3: 12, 4: 60, 5: 14, 6: 50, 7: 32}
for col_idx, w in widths.items():
    ws.column_dimensions[get_column_letter(col_idx)].width = w

# Freeze header
ws.freeze_panes = 'A6'

# ---- Footer / notes block (after data) ----
last_data_row = HEADER_ROW + len(ROWS)
notes_row = last_data_row + 2

ws.cell(row=notes_row, column=1, value='Reviewer notes').font = Font(bold=True, size=11)
ws.merge_cells(start_row=notes_row, start_column=1, end_row=notes_row, end_column=7)

reviewer_items = [
    'Confirm DCFS adaptations on A2, A3 wording, D5, E2a, E2b are correct.',
    'Decide: include the 3-question DCFS Compliance Awareness block (CA1-CA3) at the end? Default = include.',
    'Decide: survey control teams too? If yes, same set or Template 1 only (skip SPACE + DevEx)?',
    'Confirm role list now includes 5 roles: BA / Tester / Developer / Solution Architect / Testing Services Lead (cover-page demographic).',
    'Survey Monkey skip-logic constraints to verify (Robert -> Tracy).',
]
for i, item in enumerate(reviewer_items, start=1):
    ws.cell(row=notes_row + i, column=1, value=f'{i}.')
    ws.cell(row=notes_row + i, column=2, value=item)
    ws.merge_cells(start_row=notes_row + i, start_column=2, end_row=notes_row + i, end_column=7)
    ws.cell(row=notes_row + i, column=2).alignment = Alignment(wrap_text=True, vertical='top')

wb.save(OUT)
print(f'Excel created: {OUT}')
