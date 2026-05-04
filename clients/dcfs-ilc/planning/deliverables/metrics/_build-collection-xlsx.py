# pyright: reportArgumentType=false, reportOptionalMemberAccess=false
"""Build the Pilot Metrics Collection Methods Excel.

Companion to pilot-metrics-proposal-to-doit-v20260430.xlsx.
For each metric in the proposal, this Excel describes HOW to collect it:
specific steps, auto/manual, owner, cadence, prerequisites, effort,
feasibility status, and any sample query / formula.

Goal: take to the team (Matt RTE, JIRA admin, Jeff, role leads) to confirm
what we can actually collect and what's blocked.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

OUT_FULL = os.path.join(os.path.dirname(__file__), 'pilot-metrics-collection-methods-v20260430.xlsx')
OUT_LITE = os.path.join(os.path.dirname(__file__), 'pilot-metrics-collection-methods-v20260430-lite.xlsx')

# Mirror the lite filter from the proposal script — same 27 metrics.
# Apr 30 team feedback: SA Recommended swapped to SDD rejection rate; original 2 SA metrics moved to Discussion.
LITE_KEEP_NAMES = {
    # Recommended (post-Apr-30 feedback)
    'Story quality score (QUS 13-criteria)',
    'Acceptance criteria completeness',
    'SDD rejection rate',
    'Story cycle time (development time)',
    'First-pass QA yield',
    'Test coverage %',
    'Test cases per story',
    'Test automation coverage %',
    'Testing Services Playbook contributions',
    # Discussion (kept in lite)
    'Refinement cycle time',
    'Story rejection rate',
    'SDD creation time',                       # team-flagged: NEED TO REVIEW
    'SDD completeness score (pre-ARB rubric)', # team-flagged: NEED TO REVIEW
    'Pre-ARB cycle time',                      # team note: no current baseline
    'Story-to-deploy cycle time',
    'Defect density',
    'Code review cycle time',
    '% stories needing rework due to incomplete AC',
    'Negative-case coverage %',
    'Defect escape rate',
    'Test-report drafting time',
    'Deployment frequency',
    'Change failure rate',
    'Pre/mid/post AI skills delta',
    'AI tool usage volume',
    'Prompt-drift incidents',
    'SPACE survey (5 dimensions)',
}

# ---- Styling ----
HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
HEADER_FONT = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
TITLE_FONT = Font(name='Calibri', size=16, bold=True, color='1F4E79')
BODY_FONT = Font(name='Calibri', size=10)
BODY_FONT_BOLD = Font(name='Calibri', size=10, bold=True)
WRAP = Alignment(wrap_text=True, vertical='top')
WRAP_CENTER = Alignment(wrap_text=True, vertical='top', horizontal='center')
THIN = Side(border_style='thin', color='B4B4B4')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

GREEN_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
YELLOW_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
RED_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

GREY_FILL = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
FEAS_FILL = {'Confirmed': GREEN_FILL, 'Pending': YELLOW_FILL, 'Blocked': RED_FILL, 'Deferred': GREY_FILL}
AUTO_FILL = {'Automated': GREEN_FILL, 'Semi-auto': YELLOW_FILL, 'Manual': RED_FILL}

HEADERS = [
    'Role', 'Metric', 'Data Source', 'Collection Method (steps)',
    'Auto / Manual', 'Owner', 'Cadence', 'Prerequisites / Access',
    'Effort per collection', 'Feasibility', 'Sample Query / Formula', 'Notes',
]

# Tuple: Role | Metric | Data Source | Method | Auto/Manual | Owner | Cadence | Prereqs | Effort | Feasibility | Sample | Notes
ROWS = [
    # ---------- Business Analyst ----------
    ('Business Analyst', 'Story quality score (QUS 13-criteria)',
     'JIRA stories',
     'Pull all stories created in sprint via JQL; run script that scores each against the 13 QUS criteria',
     'Automated', 'AI Transformation Lead', 'Per sprint',
     'JIRA REST API token (read scope on pilot project)',
     '~5 min after script setup',
     'Pending',
     'JQL: project=ILC AND created>=startOfSprint() AND issuetype=Story',
     'Need to confirm QUS scorer library; may need to build'),
    ('Business Analyst', 'Acceptance criteria completeness',
     'JIRA stories (manual sample)',
     'Random-sample 10 stories per sprint; rate each AC block 1–5 against rubric',
     'Manual', 'Pilot Governance Lead', 'Per sprint',
     'JIRA read access; rubric defined',
     '~30 min per sprint',
     'Confirmed',
     '',
     'Rubric: clarity, testability, scope, completeness, format'),
    ('Business Analyst', 'Refinement cycle time',
     'JIRA transitions',
     'JQL: stories transitioned from "Created" to "Ready for Dev"; calculate days between',
     'Automated', 'AI Transformation Lead', 'Per sprint',
     'JIRA REST API + changelog API access',
     '~5 min after script setup',
     'Pending',
     'GET /rest/api/3/issue/{key}/changelog',
     'Confirm changelog API exposed'),
    ('Business Analyst', 'Story rejection rate',
     'JIRA workflow history',
     'Count stories that transitioned BACK from "Ready for Dev" or later to a refinement state',
     'Automated', 'AI Transformation Lead', 'Per sprint',
     'JIRA REST API + changelog API',
     '~5 min after script setup',
     'Pending',
     'Filter changelog for backwards transitions',
     ''),
    ('Business Analyst', 'SME clarification requests per story',
     'JIRA comments / labels',
     'Pull all comments on stories in sprint; flag those with @-mentions or "?" / "clarif" keywords',
     'Semi-auto', 'AI Transformation Lead', 'Per sprint',
     'JIRA REST API',
     '~10 min',
     'Pending',
     'GET /rest/api/3/issue/{key}/comment',
     'Heuristic — may want to use a label convention instead'),
    ('Business Analyst', '% sprints with stories fully laid out at PI start',
     'JIRA + PI planning records',
     'At PI start, query: of N planned sprints, how many have all stories with story-points + AC + Ready-for-Dev?',
     'Semi-auto', 'AI Transformation Lead + RTE (Matt)', 'Per PI',
     'JIRA REST API + PI plan reference (Plans / Advanced Roadmaps)',
     '~30 min per PI',
     'Pending',
     '',
     'Talk to Matt re: Advanced Roadmaps API'),

    # ---------- Solution Architect ----------
    ('Solution Architect', 'SDD rejection rate',
     'SharePoint version history + ARB tracking',
     'For each SDD, count revisions from initial submission through full ARB approval (revisions → rejections); rate = revisions per SDD',
     'Semi-auto', 'AI Transformation Lead', 'Per SDD',
     'SharePoint version history + ARB outcome tracking',
     '~10 min per SDD',
     'Pending',
     '',
     'Apr 30 team feedback: replaces "SDD revision count" — same idea, team-preferred name. Confirmed Recommended for SA.'),
    ('Solution Architect', 'SDD creation time',
     'Self-reported (Tempo or sheet)',
     'SA logs hours per SDD in Tempo or shared sheet; weekly rollup',
     'Manual', 'SA (self-reported)', 'Per SDD',
     'Tempo or shared sheet',
     '~5 min per SDD',
     'Pending',
     '',
     'Apr 30 team feedback: NEED TO REVIEW before Recommended. Confirm Tempo plugin enabled; talk to Matt.'),
    ('Solution Architect', 'SDD completeness score (pre-ARB rubric)',
     'Manual rubric on sampled SDDs',
     'Pre-ARB reviewer scores SDD against existing pre-ARB checklist; capture in Confluence',
     'Manual', 'Pre-ARB reviewer (Shyam / Kashif)', 'Per SDD',
     'Pre-ARB checklist defined',
     '~10 min per SDD',
     'Pending',
     '',
     'Apr 30 team feedback: NEED TO REVIEW before Recommended. Need rubric — derive from existing pre-ARB checklist.'),
    ('Solution Architect', 'Pre-ARB cycle time',
     'JIRA architecture-review tickets',
     'JQL on architecture-review tickets; days from "Submitted" → "Approved"',
     'Automated', 'AI Transformation Lead', 'Per SDD',
     'JIRA REST API; arch-review project key',
     '~5 min',
     'Pending',
     'JQL: project=ARCH-REVIEW AND status changed to Approved',
     'Apr 30 team feedback: "We don\'t track the duration" — no current baseline. Pilot establishes baseline from sprint 1 onward.'),
    ('Solution Architect', 'ADRs logged per sprint',
     'Confluence page-label search',
     'Search Confluence for pages labeled "adr" created in sprint window',
     'Automated', 'AI Transformation Lead', 'Per sprint',
     'Confluence API token',
     '~5 min',
     'Pending',
     'GET /rest/api/content/search?cql=label=adr',
     ''),
    ('Solution Architect', 'Maximus artifact draft time',
     'Self-reported',
     'Architect logs hours per Maximus-required artifact in shared sheet',
     'Manual', 'SA (self-reported)', 'Per artifact',
     'Shared sheet',
     '~5 min per artifact',
     'Pending',
     '',
     'Need list of Maximus-required artifacts (action item from Apr 16)'),
    ('Solution Architect', 'Cross-team dependencies surfaced',
     'JIRA epic / story link analysis',
     'Pull issue links across teams; count dependencies surfaced per sprint',
     'Automated', 'AI Transformation Lead', 'Per sprint',
     'JIRA REST API',
     '~5 min',
     'Pending',
     'GET /rest/api/3/issue/{key}?fields=issuelinks',
     ''),

    # ---------- Developer ----------
    ('Dynamics Developer', 'Story cycle time (development time)',
     'JIRA transitions',
     'JQL: stories transitioned to "In Testing" in sprint; days from "In Progress" → "In Testing"',
     'Automated', 'AI Transformation Lead', 'Per sprint',
     'JIRA REST API + changelog',
     '~5 min',
     'Pending',
     '',
     'Apr 30 team feedback: end boundary at "In Testing" handoff (not "Done") — keeps this metric dev-only. Story-to-deploy cycle time covers full delivery cycle.'),
    ('Dynamics Developer', 'Story-to-deploy cycle time',
     'JIRA + Azure DevOps deploys',
     'JIRA "Ready for Dev" timestamp + ADO deploy timestamp; calculate days',
     'Semi-auto', 'AI Transformation Lead + Release Manager', 'Per story',
     'JIRA API + ADO Pipelines API + release-tagging convention',
     '~10 min',
     'Pending',
     '',
     'Need release tagging — talk to RTE / release manager'),
    ('Dynamics Developer', 'First-pass QA yield',
     'JIRA transition history',
     'Count stories where "In QA" → "Done" without going back to "In Progress" (exact JIRA workflow status names TBD)',
     'Automated', 'AI Transformation Lead', 'Per sprint',
     'JIRA REST API + changelog + confirmed status names',
     '~5 min',
     'Pending',
     '',
     'Apr 30 team feedback: OPEN — confirm exact JIRA workflow status transitions that count as "fail" trigger before locking the calculation.'),
    ('Dynamics Developer', 'Defect density',
     'JIRA issue links',
     'Sum bugs linked to stories closed in sprint; divide by story points',
     'Automated', 'AI Transformation Lead', 'Per sprint',
     'JIRA REST API; bug-link convention',
     '~5 min',
     'Pending',
     '',
     ''),
    ('Dynamics Developer', 'Code review cycle time',
     'GitHub REST API',
     'Pull PRs; days from opened → merged',
     'Automated', 'AI Transformation Lead', 'Per PR',
     'GitHub Enterprise API token; org/repo scope',
     '~5 min',
     'Pending',
     'GET /repos/{owner}/{repo}/pulls?state=closed',
     'Tier check (Enterprise needed for some analytics)'),
    ('Dynamics Developer', 'Documentation coverage on PRs',
     'GitHub PR file analysis (script)',
     'For each merged PR, check if files in /docs/ or README* were touched',
     'Automated', 'AI Transformation Lead', 'Per sprint',
     'GitHub API + docs-path convention',
     '~5 min',
     'Pending',
     '',
     ''),
    ('Dynamics Developer', 'LOC delivered (plugins / APIs / logic apps)',
     'GitHub diff statistics',
     'Sum diff stats from merged PRs in sprint',
     'Automated', 'AI Transformation Lead', 'Per sprint',
     'GitHub API',
     '~5 min',
     'Pending',
     'GET /repos/{owner}/{repo}/stats/code_frequency',
     ''),
    ('Dynamics Developer', '% stories needing rework due to incomplete AC',
     'JIRA transition + comment labels',
     'Count stories returned to BA mid-sprint; require comment label "incomplete-ac"',
     'Semi-auto', 'AI Transformation Lead', 'Per sprint',
     'JIRA REST API + label convention',
     '~10 min',
     'Pending',
     '',
     'Need label convention — define before pilot'),

    # ---------- Senior Tester ----------
    ('Senior Tester', 'Test coverage %',
     'Zephyr Scale ↔ JIRA links',
     'Count stories with at least one linked test case / total stories in sprint (story-level coverage)',
     'Automated', 'AI Transformation Lead', 'Per sprint',
     'Zephyr Scale REST API token',
     '~5 min',
     'Pending',
     '',
     'Apr 30 team feedback: kept at story-level (not AC-level) — easier to collect, matches existing Zephyr report. AC-level coverage moved to Discussion as future enhancement.'),
    ('Senior Tester', 'Test cases per story',
     'Zephyr Scale + JIRA',
     'Count test cases linked to each story; average per story',
     'Automated', 'AI Transformation Lead', 'Per sprint',
     'Zephyr API',
     '~5 min',
     'Pending',
     '',
     ''),
    ('Senior Tester', 'Negative-case coverage %',
     'Zephyr Scale + tagging',
     'Count stories with at least one test case tagged "negative" / total stories',
     'Semi-auto', 'AI Transformation Lead', 'Per sprint',
     'Zephyr API + tagging convention',
     '~10 min',
     'Pending',
     '',
     'Need tagging convention — define before pilot'),
    ('Senior Tester', 'Test-case authoring time',
     'Self-reported + Zephyr timestamps',
     'Tester logs authoring hours; cross-check with Zephyr created/updated timestamps',
     'Manual', 'Tester (self-reported)', 'Per story',
     'Shared sheet + Zephyr timestamps',
     '~2 min per story',
     'Pending',
     '',
     ''),
    ('Senior Tester', 'Defect escape rate',
     'JIRA bug labels (pre-UAT vs post-UAT)',
     'Count post-UAT bugs / total bugs in sprint',
     'Automated', 'AI Transformation Lead', 'Per sprint',
     'JIRA API + label convention (pre-uat / post-uat)',
     '~5 min',
     'Pending',
     '',
     'Confirm bug labeling convention'),
    ('Senior Tester', 'Test-report drafting time',
     'Self-reported',
     'Tester logs hours to draft sprint / PI test report',
     'Manual', 'Tester / TSL (self-reported)', 'Per sprint / per PI',
     'Shared sheet',
     '~5 min per report',
     'Pending',
     '',
     'No incumbent process — pure upside'),

    # ---------- Shift-left testing ----------
    ('Shift-left testing (BA + Tester)', 'Test plan generated at story creation',
     'JIRA + Zephyr link at story-creation timestamp',
     'For each new story, check if a test plan was attached at story-creation time',
     'Semi-auto', 'AI Transformation Lead', 'Per sprint',
     'JIRA API + Zephyr API + story-creation convention',
     '~10 min',
     'Pending',
     '',
     'New use case — Rovo generates test plan at creation'),
    ('Shift-left testing (BA + Tester)', 'Pre-validation test coverage %',
     'Zephyr ↔ JIRA at validation-handoff timestamp',
     'Compare test coverage of stories at BA-validation-handoff vs at sprint-start',
     'Semi-auto', 'AI Transformation Lead', 'Per sprint',
     'JIRA API + Zephyr API + handoff timestamp convention',
     '~10 min',
     'Pending',
     '',
     'Pairs with above'),

    # ---------- Testing Services Lead ----------
    ('Testing Services Lead', 'Test automation coverage %',
     'Zephyr Scale automation flags',
     '% test cases marked "automated" in Zephyr (single-team for pilot)',
     'Automated', 'AI Transformation Lead', 'Per sprint',
     'Zephyr API',
     '~5 min',
     'Pending',
     '',
     'Single-team applicable; works in pilot scope'),
    ('Testing Services Lead', 'Testing Services Playbook contributions',
     'Confluence page history',
     'Count of TSL-authored updates to the cross-team testing playbook in Confluence per sprint',
     'Semi-auto', 'AI Transformation Lead', 'Per sprint',
     'Confluence API + page-author convention',
     '~5 min',
     'Pending',
     '',
     'Reusable scaling artifact — TSL pilot deliverable'),
    ('Testing Services Lead', 'Testing strategy artifacts produced',
     'SharePoint / Confluence',
     'Count of testing strategy / guideline documents authored or updated per sprint',
     'Semi-auto', 'TSL (self-reported) + AI Transformation Lead', 'Per sprint',
     'Shared sheet + SharePoint / Confluence access',
     '~10 min',
     'Pending',
     '',
     'Single-team applicable; complements playbook contributions'),
    ('Testing Services Lead', 'Cross-team coverage consistency',
     'Zephyr Scale (aggregated)',
     'Pull coverage % per team; compute variance across the 12 teams',
     'Automated', 'AI Transformation Lead', 'Monthly',
     'Zephyr API for all 12 teams',
     '~15 min',
     'Deferred',
     '',
     'DEFERRED — activates at multi-team scale; not measurable in 1-team pilot'),
    ('Testing Services Lead', 'Cross-team defect trends',
     'JIRA bug data (aggregated)',
     'Defect escape rate per team; rolling 3-month avg',
     'Automated', 'AI Transformation Lead', 'Monthly',
     'JIRA API for all 12 teams',
     '~15 min',
     'Deferred',
     '',
     'DEFERRED — activates at multi-team scale'),
    ('Testing Services Lead', 'Cross-team test-report rollup time',
     'Self-reported',
     'TSL logs hours to roll up sprint / PI reports across the 12 teams',
     'Manual', 'TSL (self-reported)', 'Per sprint / per PI',
     'Shared sheet',
     '~5 min per rollup',
     'Deferred',
     '',
     'DEFERRED — activates at multi-team scale'),
    ('Testing Services Lead', 'Cross-team pattern detection count',
     'AI tool output + manual log',
     'AI surfaces similar patterns; TSL logs distinct patterns identified',
     'Manual', 'TSL', 'Monthly',
     'AI tool access + shared log',
     '~30 min',
     'Deferred',
     '',
     'DEFERRED — activates at multi-team scale'),

    # ---------- Release Quality ----------
    ('Release Quality', 'Released-story bug return rate',
     'JIRA bug links + release labels',
     'For stories released in sprint N, count bugs filed within sprints N+1..N+3 / total released',
     'Automated', 'AI Transformation Lead', 'Per sprint (trailing)',
     'JIRA API + release labeling convention',
     '~10 min',
     'Pending',
     '',
     'TRAILING — needs 3+ sprints of data before first read'),
    ('Release Quality', 'Bug-trend slope (sprint-over-sprint)',
     'JIRA bug data (rolling window)',
     'Plot post-release bug count over rolling 3-sprint window; compute trend slope',
     'Automated', 'AI Transformation Lead', 'Per sprint (trailing)',
     'JIRA API + release labeling convention',
     '~5 min',
     'Pending',
     '',
     'Best read after 4–6 sprints of pilot data'),

    # ---------- DORA Industry Benchmarks ----------
    ('DORA Benchmarks', 'Deployment frequency',
     'Azure DevOps pipeline history',
     'Pull deploy events per team per sprint from ADO Pipelines',
     'Automated', 'AI Transformation Lead + Release Manager', 'Per sprint',
     'ADO Pipelines API access',
     '~5 min',
     'Pending',
     'GET /_apis/release/deployments',
     'DORA core metric'),
    ('DORA Benchmarks', 'Change failure rate',
     'Azure DevOps deploys + JIRA incident labels',
     'Count deploys causing incident/hotfix/rollback / total deploys',
     'Semi-auto', 'AI Transformation Lead + Release Manager', 'Per sprint',
     'ADO Pipelines API + JIRA incident-label convention',
     '~10 min',
     'Pending',
     '',
     'DORA core; needs incident labeling convention'),
    ('DORA Benchmarks', 'Rework rate',
     'JIRA transition history + GitHub PR re-open / re-merge',
     'Count stories returned to earlier state + PRs reopened or amended after merge',
     'Automated', 'AI Transformation Lead', 'Per sprint',
     'JIRA + GitHub APIs',
     '~10 min',
     'Pending',
     '',
     'DORA 2025 new metric'),
    ('DORA Benchmarks', 'AI tool ROI composite',
     'Tool admin dashboards + estimation',
     'Compute (prompts × acceptance × est. time saved per accept) per role per week',
     'Semi-auto', 'AI Transformation Lead', 'Weekly rollup',
     'Tool admin dashboards + time-saved estimation factors',
     '~20 min weekly',
     'Pending',
     '',
     'DORA 2025; estimation factors need calibration'),

    # ---------- Survey-derived ----------
    ('Survey-derived', 'Pre/mid/post AI skills delta',
     'Pilot baseline survey + mid + post',
     'Calculate average B-block skill scores per round; diff baseline → mid → post',
     'Manual', 'AI Transformation Lead', 'Pre / mid / post',
     'Survey responses (3 rounds)',
     '~30 min per round',
     'Pending',
     '',
     'Survey v0.2 already in flight (sent Apr 29)'),
    ('Survey-derived', 'Self-reported productivity perception',
     'Sprint pulse + 3-round survey',
     'Pull Likert score on productivity question; trend across rounds',
     'Manual', 'AI Transformation Lead', 'Per sprint + 3 rounds',
     'Survey tool',
     '~10 min per round',
     'Pending',
     '',
     ''),
    ('Survey-derived', 'Self-reported quality perception',
     'Sprint pulse + 3-round survey',
     'Pull Likert score on quality question; trend across rounds',
     'Manual', 'AI Transformation Lead', 'Per sprint + 3 rounds',
     'Survey tool',
     '~10 min per round',
     'Pending',
     '',
     'Pairs with measured first-pass QA yield'),
    ('Survey-derived', 'Adoption barrier index',
     '3-round survey (E-block adoption barrier questions)',
     'Aggregate E-block responses; track trend across rounds',
     'Manual', 'AI Transformation Lead', '3 rounds',
     'Survey tool',
     '~15 min per round',
     'Pending',
     '',
     'E1-E3 already capture this'),

    # ---------- Cross-cutting ----------
    ('Cross-cutting', 'AI tool usage volume',
     'GitHub Copilot / Rovo / M365 admin dashboards',
     'Weekly screenshot or CSV export from each admin dashboard',
     'Manual', 'AI Transformation Lead', 'Weekly',
     'Admin access to each tool dashboard',
     '~10 min per dashboard',
     'Pending',
     '',
     'Tier check needed — Rovo Enterprise for some breakdowns'),
    ('Cross-cutting', 'AI suggestion acceptance rate',
     'Tool admin dashboards',
     'Same dashboards; pull accept/edit/reject ratios',
     'Manual', 'AI Transformation Lead', 'Weekly',
     'Admin access',
     '~5 min',
     'Pending',
     '',
     ''),
    ('Cross-cutting', 'Prompt-drift incidents',
     'Incident reports + spot-check log',
     'Pilot members log near-misses; Governance Lead reviews weekly',
     'Manual', 'Pilot Governance Lead', 'Ongoing',
     'Teams channel + log template',
     '~15 min per week',
     'Confirmed',
     '',
     'Required by DoIT AI Policy §6 + §7'),
    ('Cross-cutting', 'SPACE survey (5 dimensions)',
     'Survey (Survey Monkey or alt)',
     'Distribute survey pre-pilot, mid-pilot, post-pilot; rollup scores per dimension',
     'Manual', 'AI Transformation Lead', 'Pre / mid / post',
     'Survey tool selected',
     '~30 min per round',
     'Pending',
     '',
     'Survey tool not yet chosen'),
    ('Cross-cutting', 'AI tool satisfaction pulse',
     'Survey',
     '5-question pulse survey at end of each sprint',
     'Manual', 'AI Transformation Lead', 'Per sprint',
     'Survey tool selected',
     '~10 min per round',
     'Pending',
     '',
     ''),
]


def style_header_row(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP_CENTER
        cell.border = BORDER


def style_body_cell(cell, bold=False, center=False):
    cell.font = BODY_FONT_BOLD if bold else BODY_FONT
    cell.alignment = WRAP_CENTER if center else WRAP
    cell.border = BORDER


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_workbook(rows_list, out_path, subtitle_extra=''):
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = 'Collection Methods'

    ws.cell(row=1, column=1,
            value='Pilot Metrics — Collection Methods (V20260430)').font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    ws.row_dimensions[1].height = 24

    subtitle = (
        'For each metric in the proposal, this sheet describes HOW it gets collected: '
        'specific steps, level of automation, owner, cadence, prerequisites (access / '
        'tokens / labels / rubrics), estimated effort, and a feasibility flag '
        '(Confirmed / Pending / Blocked / Deferred). Take to Matt (RTE), JIRA admin, and role '
        'leads to confirm what we can actually pull and what is blocked.'
        + subtitle_extra
    )
    sc = ws.cell(row=2, column=1, value=subtitle)
    sc.font = BODY_FONT
    sc.alignment = WRAP
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(HEADERS))
    ws.row_dimensions[2].height = 56

    for i, h in enumerate(HEADERS, start=1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, len(HEADERS))

    # widths: Role, Metric, Source, Method, Auto, Owner, Cadence, Prereqs, Effort, Feas, Sample, Notes
    set_widths(ws, [22, 36, 28, 50, 14, 26, 18, 32, 18, 14, 36, 30])

    for i, row in enumerate(rows_list, start=4):
        role, metric, src, method, auto, owner, cadence, prereqs, effort, feas, sample, notes = row
        values = [role, metric, src, method, auto, owner, cadence, prereqs, effort, feas, sample, notes]
        for j, val in enumerate(values, start=1):
            cell = ws.cell(row=i, column=j, value=val)
            style_body_cell(cell, bold=(j == 2), center=(j in (5, 7, 9, 10)))
            if j == 5 and auto in AUTO_FILL:
                cell.fill = AUTO_FILL[auto]
            if j == 10 and feas in FEAS_FILL:
                cell.fill = FEAS_FILL[feas]
        ws.row_dimensions[i].height = 64

    ws.freeze_panes = ws.cell(row=4, column=3)

    wb.save(out_path)
    print(f'Wrote {out_path} with {len(rows_list)} metrics')


def main():
    # Full version (internal canonical reference)
    build_workbook(ROWS, OUT_FULL)

    # Lite version (for team review) — same 27 metrics as the proposal lite
    lite_rows = [r for r in ROWS if r[1] in LITE_KEEP_NAMES]
    build_workbook(
        lite_rows,
        OUT_LITE,
        subtitle_extra=' — LITE VERSION: trimmed to ~27 metrics matching the lite proposal Excel.',
    )


if __name__ == '__main__':
    main()
