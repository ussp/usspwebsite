#!/usr/bin/env python3
# pyright: reportArgumentType=false, reportOptionalMemberAccess=false
"""Build the Pilot Metrics Proposal — broader candidate inventory for team review.

Single-sheet Excel — wide list of candidate metrics flagged with:
  - Collection Difficulty (Easy / Medium / Hard)
  - AI Win Likelihood (High / Medium / Low)
  - Velocity Impact (how it drives sprint velocity)
  - Quality Impact (how it reflects quality of work)

Team reviews and re-prioritizes based on (a) what can actually be collected
and (b) where AI is most likely to show measurable improvement.

Companion thinking doc: pilot-metrics-proposal-to-doit-v20260430.md.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

OUT_FULL = os.path.join(os.path.dirname(__file__), 'pilot-metrics-proposal-to-doit-v20260430.xlsx')
OUT_LITE = os.path.join(os.path.dirname(__file__), 'pilot-metrics-proposal-to-doit-v20260430-lite.xlsx')

# ---- Styling ----
HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
HEADER_FONT = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
TITLE_FONT = Font(name='Calibri', size=16, bold=True, color='1F4E79')
BODY_FONT = Font(name='Calibri', size=10)
BODY_FONT_BOLD = Font(name='Calibri', size=10, bold=True)
WRAP = Alignment(wrap_text=True, vertical='top')
WRAP_CENTER = Alignment(wrap_text=True, vertical='top', horizontal='center')
THIN = Side(border_style='thin', color='B4B4B4')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

GREEN_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
YELLOW_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
RED_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

DIFF_FILL = {'Easy': GREEN_FILL, 'Medium': YELLOW_FILL, 'Hard': RED_FILL}
WIN_FILL = {'High': GREEN_FILL, 'Medium': YELLOW_FILL, 'Low': RED_FILL}

HEADERS = [
    'Role', 'Metric', 'Metric Type', 'What It Measures', 'Data Source',
    'Collection Difficulty', 'AI Win Likelihood',
    'Velocity Impact (drives sprint velocity?)',
    'Quality Impact (reflects quality of work?)',
    'Recommended', 'Status', 'Team Priority', 'Notes',
]

# Metric Type — Performance proves productivity improvement; everything else is supporting context.
# Performance = quantitative work output deltas
# Adoption    = are people using the tool?
# Capability  = are people getting better at AI-assisted work?
# Sentiment   = how do people feel about AI? (self-reported)
# Process Safety = governance / compliance signal
METRIC_TYPE = {
    'Adoption': {
        'AI tool usage volume',
        'AI suggestion acceptance rate',
        'AI tool ROI composite',
    },
    'Capability': {
        'Pre/mid/post AI skills delta',
        'Adoption barrier index',
    },
    'Sentiment': {
        'Self-reported productivity perception',
        'Self-reported quality perception',
        'SPACE survey (5 dimensions)',
        'AI tool satisfaction pulse',
    },
    'Process Safety': {
        'Prompt-drift incidents',
    },
    # Everything else defaults to Performance.
}


def get_metric_type(name):
    for t, names in METRIC_TYPE.items():
        if name in names:
            return t
    return 'Performance'

# Recommended metrics for adoption per role; rest are discussion candidates.
# Apr 30 team feedback applied: SA Recommended = SDD rejection rate (others moved to Need-to-Review).
RECOMMENDED = {
    ('Business Analyst', 'Story quality score (QUS 13-criteria)'),
    ('Business Analyst', 'Acceptance criteria completeness'),
    ('Solution Architect', 'SDD rejection rate'),
    ('Dynamics Developer', 'Story cycle time (development time)'),
    ('Dynamics Developer', 'First-pass QA yield'),
    ('Senior Tester', 'Test coverage %'),
    ('Senior Tester', 'Test cases per story'),
    ('Testing Services Lead', 'Test automation coverage %'),
    ('Testing Services Lead', 'Testing Services Playbook contributions'),
}

# Metrics to keep in the LITE version (~27 total: Recommended + curated Discussion subset).
# Lite version is for team review; the full version is the internal/canonical reference.
LITE_KEEP_DISCUSSION = {
    # BA — keep speed + rework signals
    'Refinement cycle time',
    'Story rejection rate',
    # SA — Apr 30 team flagged 2 of these as "Need to Review" — keep visible in lite
    'SDD creation time',
    'SDD completeness score (pre-ARB rubric)',
    'Pre-ARB cycle time',
    # Dynamics Developer — keep DORA-like signals
    'Story-to-deploy cycle time',
    'Defect density',
    'Code review cycle time',
    '% stories needing rework due to incomplete AC',
    # Senior Tester — keep coverage depth + escape signals
    'Negative-case coverage %',
    'Defect escape rate',
    'Test-report drafting time',
    # DORA — keep core 2
    'Deployment frequency',
    'Change failure rate',
    # Survey-derived — keep skills delta as the headline survey metric
    'Pre/mid/post AI skills delta',
    # Cross-cutting — keep governance-required only
    'AI tool usage volume',
    'Prompt-drift incidents',
    'SPACE survey (5 dimensions)',
}

# Tuple: Role | Metric | What | Source | Difficulty | Win | Velocity Impact | Quality Impact | Notes
METRICS = [
    # ---------- Business Analyst ----------
    ('Business Analyst', 'Story quality score (QUS 13-criteria)',
     'Programmatic scoring of story content against 13 quality criteria',
     'JIRA stories', 'Easy', 'High',
     'Yes — better stories reduce mid-sprint rework, freeing sprint capacity',
     'Yes — direct measure of input quality',
     'Upstream lever per Apr 14 PO walkthrough (story quality is BA-skill-limited)'),
    ('Business Analyst', 'Acceptance criteria completeness',
     'Score 1–5 on AC quality for sampled stories',
     'JIRA stories (manual sample)', 'Easy', 'High',
     'Yes — complete AC means less mid-sprint clarification, faster dev',
     'Yes — direct quality check',
     'Manual rubric, ~10-story sample per sprint'),
    ('Business Analyst', 'Refinement cycle time',
     'Days from "Story created" → "Ready for Dev"',
     'JIRA transitions', 'Easy', 'Medium',
     'Yes — faster refinement = stories ready for sprint commitment sooner',
     'Indirect',
     'Aligns with the existing Preparation Metrics slide'),
    ('Business Analyst', 'Story rejection rate',
     '% of stories returned to BA after review',
     'JIRA workflow history', 'Easy', 'Medium',
     'Yes — fewer rejections = stories enter sprints faster',
     'Yes — quality gate signal',
     ''),
    ('Business Analyst', 'SME clarification requests per story',
     'Count of JIRA comments asking BA for clarification',
     'JIRA comments / labels', 'Easy', 'Medium',
     'Yes — fewer clarifications = less developer wait time',
     'Indirect (proxy for story clarity)',
     ''),
    ('Business Analyst', '% sprints with stories fully laid out at PI start',
     'Preparation readiness — what % of sprints in a PI have stories fully refined at PI start',
     'JIRA + PI planning records', 'Medium', 'High',
     'Yes — directly drives sprint commit-vs-completion',
     'Indirect (preparation quality proxy)',
     'From Apr 14 PO walkthrough — typical state is 4 of 6 sprints fully laid out post-PI'),

    # ---------- Solution Architect ----------
    ('Solution Architect', 'SDD rejection rate',
     'Number of revisions an SDD goes through from initial submission through full ARB approval (team-named metric — "rejection rate" framing)',
     'SharePoint version history + ARB tracking', 'Easy', 'Medium',
     'Yes — fewer revisions = faster ARB approval',
     'Yes — direct quality of first draft',
     'Team feedback (Apr 30): replaces "SDD revision count" — same idea, team-preferred framing. Confirmed Recommended for SA.'),
    ('Solution Architect', 'SDD creation time',
     'Hours to draft a Solution Design Document per feature',
     'Self-reported (Tempo or sheet)', 'Medium', 'High',
     'Yes — faster SDD = less waiting on architect',
     'Indirect',
     'Team feedback (Apr 30): NEED TO REVIEW — team flagged for further discussion before Recommended.'),
    ('Solution Architect', 'SDD completeness score (pre-ARB rubric)',
     'Rubric score on SDD completeness against the pre-ARB checklist',
     'Manual rubric on sampled SDDs', 'Easy', 'High',
     'Yes — complete SDDs reduce ARB rework',
     'Yes — direct quality of architecture artifact',
     'Team feedback (Apr 30): NEED TO REVIEW — team flagged for further discussion before Recommended.'),
    ('Solution Architect', 'Pre-ARB cycle time',
     'Days from SDD draft → pre-ARB approval',
     'JIRA architecture-review tickets', 'Medium', 'Medium',
     'Yes — faster review = code starts sooner',
     'Indirect',
     'Team feedback (Apr 30): "We don\'t track the duration" — no current baseline data. Pilot would establish baseline from sprint 1 onward.'),
    ('Solution Architect', 'ADRs logged per sprint',
     'Count of architecture decision records in Confluence',
     'Confluence page-label search', 'Easy', 'Medium',
     'Indirect',
     'Yes — documentation completeness',
     ''),
    ('Solution Architect', 'Maximus artifact draft time',
     'Hours to draft Maximus-required compliance artifacts (DevOps strategy, test strategy, ARB docs)',
     'Self-reported', 'Medium', 'High',
     'Indirect (frees architect time for other work)',
     'Yes — auditor compliance quality',
     'From Apr 16 architecture walkthrough — Maximus artifacts are a strong AI doc-authoring use case'),
    ('Solution Architect', 'Cross-team dependencies surfaced',
     'Count of cross-team dependencies identified per sprint',
     'JIRA epic / story link analysis', 'Easy', 'Low',
     'Yes — surfaced deps = fewer mid-sprint blockers',
     'Indirect',
     ''),

    # ---------- Developer ----------
    ('Dynamics Developer', 'Story cycle time (development time)',
     'Days from "In Progress" → "In Testing" — development-time only (handoff to testing), not full delivery cycle',
     'JIRA transitions', 'Easy', 'High',
     'Yes — direct development throughput',
     'Indirect',
     'Team feedback (Apr 30): boundary set at "In Testing" handoff (not "Done") — keeps this metric dev-only. Story-to-deploy cycle time captures full delivery cycle separately. DCFS work order: configuration-first → North52 → custom code. AI lift expected at each tier. Gated on R-22 for custom-code tier.'),
    ('Dynamics Developer', 'Story-to-deploy cycle time',
     'End-to-end days from "Ready for Dev" → deployed',
     'JIRA + Azure DevOps deploys', 'Easy', 'High',
     'Yes — end-to-end velocity',
     'Indirect',
     'Captures the full delivery lever; aligns with Completed Work slide. Gated on R-22'),
    ('Dynamics Developer', 'First-pass QA yield',
     '% of stories passing QA on first attempt',
     'JIRA transition history', 'Easy', 'High',
     'Yes — less rework = more committed stories complete',
     'Yes — direct quality signal',
     'Team feedback (Apr 30): OPEN — need to confirm exact JIRA workflow status names that count as the "fail" trigger for the calculation. Quality guardrail; proves AI isn\'t trading speed for defects.'),
    ('Dynamics Developer', 'Defect density',
     'Bugs raised per story point delivered',
     'JIRA issue links', 'Easy', 'Medium',
     'Yes — fewer defects = less debug time',
     'Yes — direct quality',
     ''),
    ('Dynamics Developer', 'Code review cycle time',
     'Hours from PR opened → PR merged',
     'GitHub REST API', 'Easy', 'High',
     'Yes — faster reviews = faster merges',
     'Indirect',
     ''),
    ('Dynamics Developer', 'Documentation coverage on PRs',
     '% of PRs that include documentation updates',
     'GitHub PR file analysis (script)', 'Easy', 'High',
     'Indirect',
     'Yes — direct documentation quality',
     ''),
    ('Dynamics Developer', 'LOC delivered (plugins / APIs / logic apps)',
     'Net lines added per sprint for code artifacts',
     'GitHub diff statistics', 'Easy', 'Low',
     'Yes — volume signal',
     'Weak (volume not quality)',
     ''),
    ('Dynamics Developer', '% stories needing rework due to incomplete AC',
     'Stories returned to BA mid-sprint due to insufficient AC',
     'JIRA transition + comment labels', 'Easy', 'Medium',
     'Yes — less mid-sprint churn',
     'Yes — quality gate signal',
     'Cross-cuts BA story-quality dependency'),

    # ---------- Senior Tester ----------
    ('Senior Tester', 'Test coverage %',
     '% of stories with at least one linked test case',
     'Zephyr Scale ↔ JIRA links', 'Easy', 'High',
     'Indirect',
     'Yes — direct quality / completeness',
     'Team feedback (Apr 30): kept at story-level (not AC-level) — easier to collect, matches existing Zephyr report. AC-level coverage moved to Discussion as a future enhancement.'),
    ('Senior Tester', 'Test cases per story',
     'Average count of test cases per JIRA story',
     'Zephyr Scale + JIRA', 'Easy', 'High',
     'Indirect',
     'Yes — coverage breadth',
     'Rovo demo (Apr 24): generated 5 cases from 1 story'),
    ('Senior Tester', 'Negative-case coverage %',
     '% of stories where at least one negative / edge / role-denial case exists',
     'Zephyr Scale + tagging', 'Medium', 'High',
     'Indirect',
     'Yes — direct quality (catches edge cases)',
     'Rovo proven to surface negative cases the team hadn\'t written'),
    ('Senior Tester', 'Test-case authoring time',
     'Hours to author test cases per story (the 20% AI-movable slice)',
     'Self-reported + Zephyr timestamps', 'Medium', 'High',
     'Yes — frees tester capacity for execution',
     'Indirect',
     '80% manual UI ceiling — execution time NOT a target'),
    ('Senior Tester', 'Defect escape rate',
     '% defects found post-UAT vs total defects',
     'JIRA bug labels (pre-UAT vs post-UAT)', 'Easy', 'Medium',
     'Indirect',
     'Yes — direct quality',
     ''),
    ('Senior Tester', 'Test-report drafting time',
     'Hours to draft sprint / PI test report',
     'Self-reported', 'Medium', 'High',
     'Indirect (frees TSL/tester capacity)',
     'Yes — visibility quality',
     'No incumbent process — pure upside (Apr 27)'),

    # ---------- Shift-left testing (NEW per Apr 30 conversation) ----------
    ('Shift-left testing (BA + Tester)', 'Test plan generated at story creation',
     '% of new stories where a draft test plan is attached at creation (before BA/PMO validation)',
     'JIRA + Zephyr link at story-creation timestamp', 'Medium', 'High',
     'Yes — testing starts earlier, less rework downstream',
     'Yes — coverage check happens upstream',
     'New use case — Rovo generates test plan at story creation, attached for BA/PMO validation'),
    ('Shift-left testing (BA + Tester)', 'Pre-validation test coverage %',
     '% of stories with linked test cases BEFORE BA/PMO validation handoff',
     'Zephyr ↔ JIRA at validation-handoff timestamp', 'Medium', 'High',
     'Yes — coverage gaps caught upstream, fewer mid-sprint surprises',
     'Yes — coverage is part of validation criteria',
     'Pairs with above; shifts coverage check left'),

    # ---------- Testing Services Lead ----------
    ('Testing Services Lead', 'Test automation coverage %',
     '% of test cases automated (Eggplant / Zephyr automated) for the pilot team',
     'Zephyr Scale automation flags', 'Easy', 'Medium',
     'Yes — automation reduces tester load per sprint',
     'Yes — repeatable quality',
     'Single-team applicable; works in pilot scope'),
    ('Testing Services Lead', 'Testing Services Playbook contributions',
     'Count of updates to the cross-team testing playbook authored by TSL per sprint',
     'Confluence page history', 'Easy', 'High',
     'Indirect (frees future-team load via reusable artifacts)',
     'Yes — playbook is the reusable scaling artifact',
     'TSL\'s pilot deliverable that gets the program ready to scale'),
    ('Testing Services Lead', 'Testing strategy artifacts produced',
     'Count of testing strategy / guideline / playbook sections authored or updated per sprint',
     'SharePoint / Confluence', 'Easy', 'Medium',
     'Indirect',
     'Yes — strategy quality and reusability',
     'Single-team applicable; complements playbook contributions'),
    ('Testing Services Lead', 'Cross-team coverage consistency',
     'Variance in test coverage across the 12 ILC teams',
     'Zephyr Scale (aggregated)', 'Easy', 'Medium',
     'Indirect',
     'Yes — portfolio quality consistency',
     'DEFERRED — activates at multi-team scale; not measurable in 1-team pilot'),
    ('Testing Services Lead', 'Cross-team defect trends',
     'Defect escape-rate trend across teams (monthly moving avg)',
     'JIRA bug data (aggregated)', 'Easy', 'Medium',
     'Indirect',
     'Yes — portfolio quality signal',
     'DEFERRED — activates at multi-team scale; not measurable in 1-team pilot'),
    ('Testing Services Lead', 'Cross-team test-report rollup time',
     'Hours to roll up sprint / PI reports across 12 teams',
     'Self-reported', 'Medium', 'High',
     'Indirect',
     'Yes — visibility quality',
     'DEFERRED — activates at multi-team scale; not measurable in 1-team pilot'),
    ('Testing Services Lead', 'Cross-team pattern detection count',
     'Count of similar work patterns across teams surfaced via AI',
     'AI tool output + manual log', 'Hard', 'Medium',
     'Yes — reuse opportunities reduce duplicate work',
     'Yes — best-of-breed propagation',
     'DEFERRED — activates at multi-team scale; not measurable in 1-team pilot'),

    # ---------- Release Quality (trailing — needs multi-sprint trend) ----------
    ('Release Quality', 'Released-story bug return rate',
     '% of released stories that have a bug filed against them within N sprints (default N=3)',
     'JIRA bug links + release labels', 'Medium', 'Medium',
     'Yes — fewer post-release bugs = less rework, more capacity for new work',
     'Yes — direct release quality',
     'TRAILING INDICATOR — needs at least 2–3 sprints to see trend. Requires release tagging convention. Pairs with defect escape rate but operates on a longer horizon.'),
    ('Release Quality', 'Bug-trend slope (sprint-over-sprint)',
     'Direction of post-release bug count (rising / flat / falling) over rolling 3-sprint window',
     'JIRA bug data (rolling window)', 'Medium', 'Medium',
     'Indirect (signals whether AI-assisted work is improving release stability)',
     'Yes — quality trend over time',
     'Best read after 4–6 sprints of pilot data'),

    # ---------- DORA Industry Benchmarks (per DORA 2025 AI-Assisted SW Dev report) ----------
    ('DORA Benchmarks', 'Deployment frequency',
     'Number of deployments to test/UAT/Prod per team per sprint',
     'Azure DevOps pipeline history', 'Easy', 'Medium',
     'Yes — direct sprint throughput signal',
     'Indirect',
     'DORA core metric. Easy from ADO. Note: DORA 2025 found AI gains often DON\'T move this at the team level (Productivity Paradox).'),
    ('DORA Benchmarks', 'Change failure rate',
     '% of deploys that cause a production incident, hotfix, or rollback',
     'Azure DevOps deploys + JIRA incident labels', 'Medium', 'High',
     'Yes — fewer failures = less reactive work, more sprint capacity',
     'Yes — direct release stability signal',
     'DORA core metric. DORA 2025 specifically flagged that AI adoption can hurt this — important guardrail.'),
    ('DORA Benchmarks', 'Rework rate',
     '% of stories or PRs redone within the sprint or shortly after',
     'JIRA transition history + GitHub PR re-open / re-merge', 'Medium', 'High',
     'Yes — less rework = more committed work completes',
     'Yes — direct quality signal',
     'DORA 2025 new metric. Catches AI-induced quality slippage early.'),
    ('DORA Benchmarks', 'AI tool ROI composite',
     'Composite of (usage × acceptance × estimated time saved per accepted suggestion)',
     'Tool admin dashboards + estimation', 'Medium', 'High',
     'Indirect (signals whether AI investment is paying off)',
     'Indirect',
     'DORA 2025 new metric. Useful for executive comms — single ROI number.'),

    # ---------- Survey-derived (pilot baseline / mid / post + sprint pulse) ----------
    ('Survey-derived', 'Pre/mid/post AI skills delta',
     'Average lift across B-block skill questions between baseline / mid-pilot / post-pilot',
     'Pilot baseline survey + mid + post', 'Medium', 'High',
     'Indirect (skill growth → sustained productivity)',
     'Yes — capability quality',
     'From the v0.2 baseline survey already in flight (sent Apr 29 for review)'),
    ('Survey-derived', 'Self-reported productivity perception',
     'Likert score on "AI tools have improved my productivity" — tracked across rounds',
     'Sprint pulse + 3-round survey', 'Easy', 'Medium',
     'Adoption + sentiment signal',
     'Indirect',
     'Subjective but a strong leading indicator of sustained adoption'),
    ('Survey-derived', 'Self-reported quality perception',
     'Likert score on "AI tools have improved the quality of my work"',
     'Sprint pulse + 3-round survey', 'Easy', 'Medium',
     'Indirect',
     'Yes — perceived quality lift signal',
     'Pairs with measured first-pass QA yield + defect density'),
    ('Survey-derived', 'Adoption barrier index',
     'Count / weight of adoption barriers reported across rounds — trending down over pilot',
     '3-round survey (E-block adoption barrier questions)', 'Easy', 'Medium',
     'Indirect (lower barriers = wider adoption)',
     'Indirect',
     'Survey E1-E3 already capture this; trend across rounds is the metric'),

    # ---------- Cross-cutting (all roles) ----------
    ('Cross-cutting', 'AI tool usage volume',
     'Prompts sent, hours of tool use per role per week',
     'GitHub Copilot / Rovo / M365 admin dashboards', 'Easy', 'n/a',
     'Adoption signal',
     'Adoption signal',
     'Required for DoIT MEASURE / MONITOR oversight'),
    ('Cross-cutting', 'AI suggestion acceptance rate',
     '% of AI suggestions accepted / edited / rejected',
     'Tool admin dashboards', 'Easy', 'n/a',
     'Adoption + tuning signal',
     'Quality guardrail',
     ''),
    ('Cross-cutting', 'Prompt-drift incidents',
     'Cases where PII / case data / out-of-scope content surfaced in a prompt',
     'Incident reports + spot-check log', 'Medium', 'n/a',
     'Risk',
     'Yes — required quality / safety control',
     'Required by DoIT AI Policy §6 (HITL) + §7 (monitoring)'),
    ('Cross-cutting', 'SPACE survey (5 dimensions)',
     'Team health — Satisfaction, Performance, Activity, Communication, Efficiency',
     'Survey (Survey Monkey or alt)', 'Medium', 'n/a',
     'Capacity / sustainability signal',
     'Captures qualitative quality gains',
     '3 rounds per pilot'),
    ('Cross-cutting', 'AI tool satisfaction pulse',
     '5-question satisfaction pulse per sprint',
     'Survey', 'Easy', 'n/a',
     'Adoption signal',
     'Indirect',
     ''),
]


def style_header_row(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP_CENTER
        cell.border = BORDER


def style_body_cell(cell, bold=False, center=False):
    cell.font = BODY_FONT_BOLD if bold else BODY_FONT
    cell.alignment = WRAP_CENTER if center else WRAP
    cell.border = BORDER


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_workbook(metrics_list, out_path, subtitle_extra=''):
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = 'Metrics'

    ws.cell(row=1, column=1, value='Pilot Metrics Proposal — DCFS AI Pilot (V20260430)').font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    ws.row_dimensions[1].height = 24

    subtitle = (
        'Broader candidate inventory for team review. Each row flagged with '
        'Collection Difficulty (Easy / Medium / Hard), AI Win Likelihood (High / Medium / Low), '
        'Velocity Impact (drives sprint throughput?), and Quality Impact (reflects quality of work?). '
        'Team reviews and picks priorities based on (a) what can actually be collected, '
        '(b) where AI is most likely to show measurable improvement, and (c) coverage of both '
        'velocity and quality dimensions. Aim is to lock 8–12 metrics in design phase per Apr 24 D-3.'
        + subtitle_extra
    )
    sc = ws.cell(row=2, column=1, value=subtitle)
    sc.font = BODY_FONT
    sc.alignment = WRAP
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(HEADERS))
    ws.row_dimensions[2].height = 56

    for i, h in enumerate(HEADERS, start=1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, len(HEADERS))

    # widths: Role, Metric, Type, What, Source, Diff, Win, Velocity, Quality, Recommended, Status, Priority, Notes
    set_widths(ws, [22, 36, 16, 38, 28, 14, 14, 38, 38, 16, 14, 16, 32])

    # Sort: ALL recommended metrics at top globally, then all discussion metrics.
    # Within each section, preserve original role grouping order.
    role_order = []
    for m in metrics_list:
        if m[0] not in role_order:
            role_order.append(m[0])
    rec_sorted = []
    disc_sorted = []
    for role in role_order:
        in_role = [m for m in metrics_list if m[0] == role]
        rec_sorted.extend([m for m in in_role if (m[0], m[1]) in RECOMMENDED])
        disc_sorted.extend([m for m in in_role if (m[0], m[1]) not in RECOMMENDED])
    sorted_metrics = rec_sorted + disc_sorted

    REC_FILL = PatternFill(start_color='D9EAD3', end_color='D9EAD3', fill_type='solid')

    # Metric Type colors — Performance is the headline signal, others are supporting context.
    TYPE_FILL = {
        'Performance': PatternFill(start_color='B7DEE8', end_color='B7DEE8', fill_type='solid'),  # blue
        'Adoption': PatternFill(start_color='F4CCCC', end_color='F4CCCC', fill_type='solid'),     # pink
        'Capability': PatternFill(start_color='D9D2E9', end_color='D9D2E9', fill_type='solid'),   # purple
        'Sentiment': PatternFill(start_color='FCE5CD', end_color='FCE5CD', fill_type='solid'),    # peach
        'Process Safety': PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'),  # cream
    }

    for i, row in enumerate(sorted_metrics, start=4):
        role, metric, what, src, diff, win, velocity, quality, notes = row
        is_recommended = (role, metric) in RECOMMENDED
        rec_label = 'Recommended (Top 2)' if is_recommended else 'Discussion'
        mtype = get_metric_type(metric)
        values = [role, metric, mtype, what, src, diff, win, velocity, quality, rec_label, 'Proposed', '', notes]
        for j, val in enumerate(values, start=1):
            cell = ws.cell(row=i, column=j, value=val)
            style_body_cell(cell, bold=(j in (2, 10) if is_recommended else j == 2),
                            center=(j in (3, 6, 7, 10, 11, 12)))
            if j == 3 and mtype in TYPE_FILL:
                cell.fill = TYPE_FILL[mtype]
            elif j == 6 and diff in DIFF_FILL:
                cell.fill = DIFF_FILL[diff]
            elif j == 7 and win in WIN_FILL:
                cell.fill = WIN_FILL[win]
            elif is_recommended and j not in (3, 6, 7):
                cell.fill = REC_FILL
        ws.row_dimensions[i].height = 56

    ws.freeze_panes = ws.cell(row=4, column=3)

    # Hide AI Win Likelihood column (now column G = index 7) — internal-only signal
    ws.column_dimensions['G'].hidden = True

    wb.save(out_path)
    print(f'Wrote {out_path} with {len(metrics_list)} metrics')


def main():
    # Full version (internal canonical reference) — all 51 metrics
    build_workbook(METRICS, OUT_FULL)

    # Lite version (for team review) — Recommended Top 10 + curated Discussion subset (~27 total)
    lite_metrics = [
        m for m in METRICS
        if (m[0], m[1]) in RECOMMENDED or m[1] in LITE_KEEP_DISCUSSION
    ]
    build_workbook(
        lite_metrics,
        OUT_LITE,
        subtitle_extra=' — LITE VERSION: trimmed to ~27 metrics for team review (Top 10 Recommended + curated Discussion subset).',
    )


if __name__ == '__main__':
    main()
